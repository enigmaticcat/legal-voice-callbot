"""
Sinh file Excel MOS — mỗi dòng 1 câu hỏi, 4 conditions × 3 tiêu chí.

Input : evaluation/synthetic_qa.jsonl  (lấy id + question)
        evaluation/mos_mapping.json    (nếu có, để map clip file)
Output: evaluation/mos_eval_sheets.xlsx

Cấu trúc sheet MOS_Scores:
  # | Query ID | Câu hỏi | chunk_20 [N/P/C] | chunk_40 [N/P/C] | chunk_80 [N/P/C] | full [N/P/C]

Chạy:
    python scripts/generate_mos_sheets.py
"""

import json
import random
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise SystemExit("pip install openpyxl")

SYNTHETIC_QA  = Path('evaluation/synthetic_qa.jsonl')
OUT_PATH      = Path('evaluation/mos_eval_sheets.xlsx')
N_QUERIES     = 250
SEED          = 42

CONDITIONS    = ['chunk_20', 'chunk_40', 'chunk_80', 'full']
CRITERIA      = ['N', 'P', 'C']   # Naturalness, Prosody, Continuity

# Colors
HDR_COND  = {
    'chunk_20': 'C00000',  # đỏ đậm
    'chunk_40': 'ED7D31',  # cam
    'chunk_80': '375623',  # xanh lá đậm
    'full':     '1F4E79',  # xanh navy
}
ALT_FILL    = PatternFill('solid', fgColor='EEF3FB')
SCORE_FILLS = {
    'chunk_20': PatternFill('solid', fgColor='FCE4D6'),
    'chunk_40': PatternFill('solid', fgColor='FCE4D6'),
    'chunk_80': PatternFill('solid', fgColor='E2EFDA'),
    'full':     PatternFill('solid', fgColor='DDEBF7'),
}
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
THIN   = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def make_instructions_sheet(wb):
    ws = wb.create_sheet('Instructions', 0)
    ws.column_dimensions['A'].width = 110
    lines = [
        ('HƯỚNG DẪN ĐÁNH GIÁ CHẤT LƯỢNG GIỌNG ĐỌC TTS', True, 14),
        ('', False, 11),
        ('Mỗi dòng là 1 câu hỏi dinh dưỡng, được đọc bằng 4 cấu hình TTS khác nhau (chunk_20 / chunk_40 / chunk_80 / full).', False, 11),
        ('Nghe lần lượt 4 clip của cùng 1 câu, rồi chấm điểm 3 tiêu chí (N / P / C) cho mỗi clip.', False, 11),
        ('', False, 11),
        ('N — NATURALNESS (Tự nhiên): ngữ điệu tổng thể nghe có tự nhiên như người thật không', True, 11),
        ('  5 = Hoàn toàn tự nhiên   4 = Tự nhiên, đôi chỗ hơi máy   3 = Chấp nhận được', False, 11),
        ('  2 = Khó nghe, không tự nhiên nhiều   1 = Rất khó nghe', False, 11),
        ('', False, 11),
        ('P — PROSODY (Ngữ điệu): nhấn trọng âm, lên xuống giọng có hợp lý không', True, 11),
        ('  5 = Rất tự nhiên   4 = Tốt   3 = Trung bình   2 = Sai nhiều   1 = Rất sai', False, 11),
        ('', False, 11),
        ('C — CONTINUITY (Liền mạch): các phần audio có nối liền nhau trơn tru không', True, 11),
        ('  5 = Hoàn toàn liền mạch   4 = Liền mạch, 1-2 chỗ dừng nhỏ   3 = Vài chỗ ngắt', False, 11),
        ('  2 = Ngắt nhiều   1 = Giật cục, không thể nghe liên tục', False, 11),
        ('', False, 11),
        ('LƯU Ý:', True, 11),
        ('  • Nghe lại tối đa 2 lần trước khi chấm', False, 11),
        ('  • 3 dòng đầu là practice — không tính điểm', False, 11),
        ('  • Yêu cầu: tai nghe, phòng yên tĩnh, tiếng Việt bản ngữ', False, 11),
        ('  • Nhập số nguyên 1–5 vào các ô màu', False, 11),
    ]
    for text, bold, size in lines:
        ws.append([text])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=bold, size=size)
        ws.row_dimensions[row].height = 20 if text else 8


def make_scores_sheet(wb, queries):
    ws = wb.create_sheet('MOS_Scores')

    # ── Row 1: condition group headers ────────────────────────────────────────
    ws.cell(1, 1).value = '#'
    ws.cell(1, 2).value = 'Query ID'
    ws.cell(1, 3).value = 'Câu hỏi'

    col = 4
    for cond in CONDITIONS:
        # Merge 3 ô cho tên condition
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
        c = ws.cell(1, col)
        c.value     = cond
        c.font      = Font(bold=True, color='FFFFFF', size=11)
        c.fill      = PatternFill('solid', fgColor=HDR_COND[cond])
        c.alignment = CENTER
        c.border    = BORDER
        for dc in range(3):
            ws.cell(1, col+dc).border = BORDER
        col += 3

    # ── Row 2: criteria sub-headers ───────────────────────────────────────────
    ws.cell(2, 1).value = ''
    ws.cell(2, 2).value = ''
    ws.cell(2, 3).value = ''
    col = 4
    for cond in CONDITIONS:
        for crit in CRITERIA:
            c = ws.cell(2, col)
            c.value     = crit
            c.font      = Font(bold=True, color='FFFFFF', size=10)
            c.fill      = PatternFill('solid', fgColor=HDR_COND[cond])
            c.alignment = CENTER
            c.border    = BORDER
            col += 1

    # Style fixed cols row 1-2
    for row in (1, 2):
        for col_i in (1, 2, 3):
            c = ws.cell(row, col_i)
            c.font      = Font(bold=True, color='FFFFFF', size=11)
            c.fill      = PatternFill('solid', fgColor='2F5496')
            c.alignment = CENTER
            c.border    = BORDER

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 55
    for col_i in range(4, 4 + len(CONDITIONS) * 3):
        ws.column_dimensions[get_column_letter(col_i)].width = 6

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    # ── Data validation 1–5 cho tất cả ô điểm ────────────────────────────────
    score_start_col = 4
    score_end_col   = 3 + len(CONDITIONS) * 3
    score_end_row   = 2 + len(queries)
    dv = DataValidation(
        type='whole', operator='between',
        formula1='1', formula2='5',
        showErrorMessage=True,
        error='Nhập 1–5', errorTitle='Giá trị không hợp lệ',
    )
    dv.sqref = (
        f'{get_column_letter(score_start_col)}3:'
        f'{get_column_letter(score_end_col)}{score_end_row}'
    )
    ws.add_data_validation(dv)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for idx, q in enumerate(queries, 1):
        row_num = idx + 2
        alt     = idx % 2 == 0

        ws.cell(row_num, 1).value = idx
        ws.cell(row_num, 2).value = q['id']
        ws.cell(row_num, 3).value = q['question']

        for col_i in (1, 2, 3):
            c = ws.cell(row_num, col_i)
            c.border    = BORDER
            c.alignment = LEFT if col_i == 3 else CENTER
            if alt:
                c.fill = ALT_FILL

        col = 4
        for cond in CONDITIONS:
            for _ in CRITERIA:
                c = ws.cell(row_num, col)
                c.fill      = SCORE_FILLS[cond]
                c.border    = BORDER
                c.alignment = CENTER
                col += 1

        ws.row_dimensions[row_num].height = 28

    ws.freeze_panes = 'A3'


def main():
    # Load queries
    with open(SYNTHETIC_QA, encoding='utf-8') as f:
        records = [json.loads(l) for l in f if l.strip()]
    print(f'Loaded {len(records)} records from {SYNTHETIC_QA}')

    # Stratified sample 250
    by_source = {}
    for r in records:
        by_source.setdefault(r['source'], []).append(r)

    SAMPLE_COUNTS = {
        'benhvienthucuc': 78,
        'viendinhduong':  72,
        'suckhoedoisong': 51,
        'vinmec':         49,
    }
    assert sum(SAMPLE_COUNTS.values()) == N_QUERIES

    random.seed(SEED)
    selected = []
    for src, n in SAMPLE_COUNTS.items():
        selected.extend(random.sample(by_source[src], n))
    random.shuffle(selected)
    print(f'Selected {len(selected)} queries (stratified, seed={SEED})')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    make_instructions_sheet(wb)
    make_scores_sheet(wb, selected)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))
    print(f'Saved: {OUT_PATH}')
    print(f'  {N_QUERIES} rows × {len(CONDITIONS) * len(CRITERIA)} score columns '
          f'({len(CONDITIONS)} conditions × {len(CRITERIA)} criteria)')


if __name__ == '__main__':
    main()
